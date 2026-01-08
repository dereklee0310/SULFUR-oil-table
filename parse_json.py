"""
Parse ./tmp/data.json and dump result into oils.xlsx.
"""

import json
import re
from pathlib import Path
from pprint import pprint
import sys
from collections import defaultdict

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

from utils.utils import setup_logger, parse_json_args

DATA_PATH = "./tmp/data.json"
OIL_NAME_REGEX = re.compile(r"Enchantment_(.*)Oil")
EXCEL_OUTPUT_PATH = "./oils.xlsx"

COLUMN_NAME_MAPPING = {
    "displayName": "Name",
    "includedInDemo": "Demo",
    "includedInEarlyAccess": "Early Access",
    "basePrice": "Base Price",
    "CostsDurability": "Costs Durability",
    "Kick": "Recoil",
    "Reload Speed": "Reload Speed",
    "Damage": "Damage",
    "Critical damage chance": "Crit Chance",
    "Disables aiming": "Disables Aiming",
    "Projectile drag multiplier": "More Drag",
    "Spread": "Spread",
    "Loot Chance Multiplier": "Loot\nChance\nMultiplier",
    "Bullet bounces": "Bullet Bounces",
    "Time scale": "Bullet Speed",
    "Damage%": "Damage%",
    "Bullet size": "Bullet Size",
    "Bullet drop": "Bullet Drop",
    "No money drops": "No Money Drops",
    "Move speed": "Move Speed",
    "Rounds per minute": "RPM",
    "Bullet Penetration": "Bullet Penetration",
    "Jump power": "Jump Power",
    "Max Durability": "Max Durability",
    "Number of projectiles": "Projectile Amount",
    "Projectile bounciness": "Projectile Bounciness",
    "Chance this consumes ammo": "Ammo Consume Chance",
    "Chance to consume extra ammo": "Chance to Consume Extra Ammo",
    "No organs drop": "No Organs Drop",
    "Durability Per Shot": "Durability Per Shot",
    "Projectile force multiplier": "Projectile Force Multiplier",
    "Accuracy when moving": "Accuracy When Moving",
    "Enchantment Random Oil": "Enchantment Random Oil",
    "MISC": "MISC",  # Not a built-in one, for sheet name conversion
}

args = parse_json_args()
logger = setup_logger(args.logging_level)
cnt = 0

def build_oil_object(data, oil_id):
    # Enchantment_*Oil
    oil_data = data[oil_id]
    global cnt
    cnt += 1
    logger.info(f"Parsing {cnt:>3} %s", oil_data["displayName"])
    result = {
        "displayName": oil_data["displayName"],
        "includedInDemo": oil_data["includedInDemo"] ,
        "includedInEarlyAccess": oil_data["includedInEarlyAccess"],
        "basePrice": oil_data["basePrice"],
        **get_oil_definition(data, str(oil_data["appliesEnchantment"]["m_PathID"])),
    }
    return result


def get_oil_definition(data, oil_definition_id):
    # EnchantmentDefinition_*Oil
    definition_data = data[oil_definition_id]
    return {
        "CostsDurability": definition_data["CostsDurability"],
        **get_modifiers_definition(data, definition_data["modifiersApplied"]),
    }


def get_modifiers_definition(data, modifiers):
    results = {}
    for modifier in modifiers:
        # itemDescriptionName is not reliable, use label instead
        name = data[str(modifier["attribute"]["m_PathID"])]["label"]
        # 100: boolean/add, 200: multiplier, 300: bullet size
        mod_type = modifier["modType"]
        value = modifier["value"]
        if name == "Damage" and mod_type == 200:
            results["Damage%"] = value
        else:
            results[name] = value
    return results


def get_oil_types(info):
    types = []
    checks = (
        ("Kick", lambda v: v < 0),
        ("Reload Speed", lambda v: v > 0),
        ("Damage", lambda v: v > 0),
        ("Critical damage chance", lambda v: v > 0),
        ("Spread", lambda v: v < 0),
        ("Bullet bounces", lambda v: v > 0),
        ("Time scale", lambda v: v > 0),
        ("Damage%", lambda v: v > 0),
        ("Bullet size", lambda v: v > 0),
        ("Rounds per minute", lambda v: v > 0),
        ("Bullet Penetration", lambda v: v > 0),
        ("Max Durability", lambda v: v > 0),
        ("Number of projectiles", lambda v: v > 0),
        ("Chance this consumes ammo", lambda v: v < 0),
    )
    types = [key for key, cond in checks if cond(info.get(key, 0))]
    return types if types else ["MISC"]

def parse_json():
    logger.info("Parsing json file: %s", DATA_PATH)
    try:
        with open(DATA_PATH, "r", encoding="utf8") as f:
            data = json.load(f)
    except FileNotFoundError:
        logger.critical("%s not found! Please parse the game bundles first.")
        sys.exit()

    oil_infos = [build_oil_object(data, oil_id) for oil_id in data["oil_ids"]]

    oil_groups = defaultdict(list)
    for oil_info in oil_infos:
        for type in get_oil_types(oil_info):
            oil_groups[type].append(oil_info)

    # Beautify
    with pd.ExcelWriter("oils.xlsx", engine="openpyxl") as writer:
        df = pd.DataFrame(oil_infos)
        df = df.rename(columns=COLUMN_NAME_MAPPING)
        df = df.map(
            lambda x: float(f"{x:.2f}") if isinstance(x, float) and not pd.isna(x) else x
        )
        df.to_excel(writer, sheet_name="Comparison Table", index=False)
        for group_name, oil_infos in oil_groups.items():
            df = pd.DataFrame(oil_infos)
            df = df.rename(columns=COLUMN_NAME_MAPPING)
            df = df.map(
                lambda x: float(f"{x:.2f}") if isinstance(x, float) and not pd.isna(x) else x
            )
            df.to_excel(writer, sheet_name=COLUMN_NAME_MAPPING[group_name], index=False)

    # Adjust width
    wb = openpyxl.load_workbook(EXCEL_OUTPUT_PATH)
    for worksheet in wb.sheetnames:
        ws = wb[worksheet]
        dim_holder = DimensionHolder(worksheet=ws)
        for col in range(ws.min_column, ws.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(
                ws, min=col, max=col, width=20
            )
        ws.column_dimensions = dim_holder
    wb.save(EXCEL_OUTPUT_PATH)


if __name__ == "__main__":
    parse_json()